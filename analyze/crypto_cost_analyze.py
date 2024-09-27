import time
import glob
import operator
from itertools import islice
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import numpy as np
import sys
import xlsxwriter
from xlsxwriter.utility import * #xl_rowcol_to_cell
from openpyxl import load_workbook
import configparser

from time_util import *

def wazir_base(user): 
	return 'DATA/DAILY-CRYPTO/wazirx-'+user+'/WazirX_TradeReport_'

def binance_base(user): 
	return 'DATA/DAILY-CRYPTO/binance-'+user+'/'

def coindcx_base(user): 
	return 'DATA/DAILY-CRYPTO/coindcx-'+user+'/'

# ---------------------------------------------------------------------

def get_coindcx_files(base):
    fname = base + '202*.csv*'
    files = glob.glob(fname)
    if files == []: 
        print("get_coindcx_files: %s No files found %s" % (base, fname))
        return ""
    
    filetuples = []
    # file name is base_startdate_enddate
    for file in files:
    	off = len(base)
    	width = 10
    	end = file[off:off+width]
    	filetuples.append((file, end))

    return [x[0] for x in sorted(filetuples, key=operator.itemgetter(1), reverse=True)]
 
def parse_coindcx_trades(file):
	print('\nParsing file: %s ...' % file)

	df = pd.read_csv(file)

	df = df[(df.Status == 'filled')]
	df['Volume'] = df['Total Quantity']
	df['Total'] = df['Volume'] * df['Price Per Unit']
	df = df[['Market', 'Total', 'Volume']]
	#print(df)

	mkt = df.groupby('Market').agg(np.sum)
	mkt['Total'] = mkt['Total']/1000
	mkt['Average'] = mkt['Total']/mkt['Volume']	
	print(mkt)

	tot = mkt['Total'].sum()
	print('\nTotal: %.5f lacs' % (tot/100))

	return mkt

# ---------------------------------------------------------------------

def get_wazir_files(base):
    fname = base + '202*.xls*'
    files = glob.glob(fname)
    if files == []: 
        print("get_wazir_files: %s No files found %s" % (base, fname))
        return ""
    
    filetuples = []
    # file name is base_startdate_enddate
    for file in files:
    	off = len(base)
    	width = 10
    	start = file[off:off+width]
    	off += width + 1
    	end = file[off:off+width]
    	filetuples.append((file, end))

    return [x[0] for x in sorted(filetuples, key=operator.itemgetter(1), reverse=True)]
 
def parse_trades(file):
	print('\nParsing file: %s ...' % file)
	# Load in the workbook
	wb = load_workbook(file)
	#print(wb.sheetnames) 
	sheet = wb['Exchange Trades']
	print('%d rows, %d columns' % (sheet.max_row, sheet.max_column))

	#df = pd.read_excel(file, sheet_name = 1, skiprows=2)
	data = sheet.values
	cols = next(data)[1:]
	data = (islice(r, 1, None) for r in data)

	df = pd.DataFrame(data, columns=cols)
	if not 'Total' in df: df['Total'] = df['Total (Price x Volume)']
	df = df[['Market', 'Total', 'Volume']]
	#print(df)

	mkt = df.groupby('Market').agg(np.sum)
	mkt['Total'] = mkt['Total']/1000
	mkt['Average'] = mkt['Total']/mkt['Volume']	
	print(mkt)

	tot = mkt['Total'].sum()
	print('\nTotal: %.5f lacs' % (tot/100))

	return mkt

# ---------------------------------------------------------------------

import zipfile

def get_binance_p2p_files(base):
    fname = base + '202*.zip'
    files = glob.glob(fname)
    if files == []: 
        print("get_binance_files: %s No files found %s" % (base, fname))
        return ""
    
    filetuples = []
    # file name is base_startdate_enddate
    for file in files:
    	off = len(base)
    	width = 8
    	end = file[off:off+width]
    	filetuples.append((file, end))

    return [x[0] for x in sorted(filetuples, key=operator.itemgetter(1))]

def get_binance_trade_files(base):
    fname = base + '202*.xlsx'
    files = glob.glob(fname)
    if files == []: 
        print("get_binance_trade_files: %s No files found %s" % (base, fname))
        return ""
    
    filetuples = []
    # file name is base_startdate_enddate
    for file in files:
    	off = len(base)
    	width = 10
    	end = file[off:off+width]
    	filetuples.append((file, end))

    return [x[0] for x in sorted(filetuples, key=operator.itemgetter(1), reverse=True)]

def parse_binance_trades(file):
	print('\nParsing file: %s ...' % file)
	# Load in the workbook
	wb = load_workbook(file)
	#print(wb.sheetnames) 
	sheet = wb['sheet1']
	#print('%d rows, %d columns' % (sheet.max_row, sheet.max_column))

	#df = pd.read_excel(file, sheet_name = 1, skiprows=2)
	data = sheet.values
	cols = next(data)[1:]
	data = (islice(r, 1, None) for r in data)

	df = pd.DataFrame(data, columns=cols)
	#print(df)

	df['Fee'] = df['Fee'].astype(float)
	fee = df.groupby('Fee Coin').agg(np.sum)

	df = df[['Market', 'Amount', 'Total']]
	df['Volume'] = df['Amount'].astype(float)
	df['Total'] = df['Total'].astype(float)

	mkt = df.groupby('Market').agg(np.sum)
	print(mkt)
	tot = mkt['Total'].sum()
	print('\nTotal: %.5f lacs' % (tot/100))
	return mkt, fee

def parse_binance_p2p(file, process_trades):
	print('\nParsing file: %s (process trades %s) ...' % (file, process_trades))

	odir = file.replace('.zip', '')
	ofile = file.replace('.zip', '.csv')
	with zipfile.ZipFile(file, 'r') as zip_ref:
	    zip_ref.extractall(odir)

	files = glob.glob(odir+'/part-*.csv')
	if len(files) != 1: 
		print('No unique csv file after unzipping (%d files)' % len(files))
		print(files)
		return

	df = pd.read_csv(files[0])
	df = df[df['Status'] == 'Completed']
	df['Market']=df['Asset Type']+'INR'
	df.rename(columns={'Quantity':'Volume', 'Total Price':'Total'}, inplace=True)
	df = df[['Market', 'Total', 'Volume']]

	#print(df)

	mkt = df.groupby('Market').agg(np.sum)
	mkt['Total'] = mkt['Total']/1000
	print(mkt)

	if process_trades:
		tfiles = get_binance_trade_files(binance_base(user)+'Export Trade History-')
		if len(tfiles)>0:
			trades, fee = parse_binance_trades(tfiles[0])
			solfee = fee.loc[['SOL']]['Fee'][0]
			vol = trades.loc[['SOLBNB']]['Volume'][0]
			tot = mkt.loc[['BNBINR']]['Total'][0]
			newrow = pd.DataFrame({'Total':tot, 'Volume':vol-solfee}, index=['SOLINR'])
			newrow.index.name = 'Market'
			mkt = mkt.append(newrow)
			mkt = mkt.drop('BNBINR')

	mkt['Average'] = mkt['Total']/mkt['Volume']	
	print(mkt)
	tot = mkt['Total'].sum()
	print('\nTotal: %.5f lacs' % (tot/100))
	return mkt

# ---------------------------------------------------------------------

print('='*80)

now = time.time()
nowstr = time.strftime('%Y-%m-%d', time.gmtime(now))

config = configparser.ConfigParser()
config.read('config.ini')

data = config['data']
basecryptocostdet = data['basecryptocostdet']

df = {}
tdf = pd.DataFrame()

for user in ['sandesh']:
	print('\n******** Analyzing COINDCX %s ...' % user)
	files = get_coindcx_files(coindcx_base(user))
	df[user] = pd.DataFrame()
	for file in files:
		dftemp = parse_coindcx_trades(file)
		dftemp['Owner'] = user
		dftemp['Held in'] = 'coindcx'
		df[user] = df[user].append(dftemp)
	tdf = tdf.append(df[user])

for user in ['sandesh']:
	print('\n******** Analyzing BINANCE %s ...' % user)
	files = get_binance_p2p_files(binance_base(user))
	df[user] = pd.DataFrame()
	for i in range(len(files)):
		dftemp = parse_binance_p2p(files[i], True if (i==0) else False)
		dftemp['Owner'] = user
		dftemp['Held in'] = 'binance'
		df[user] = df[user].append(dftemp)
	tdf = tdf.append(df[user])

for user in ['skgoel', 'anshu', 'sandesh']:
	print('\n******** Analyzing WAZIRX %s ...' % user)
	files = get_wazir_files(wazir_base(user))
	df[user] = pd.DataFrame()
	for file in files:
		dftemp = parse_trades(file)
		dftemp['Owner'] = user
		dftemp['Held in'] = 'wazirx'
		df[user] = df[user].append(dftemp)
	tdf = tdf.append(df[user])

print('\n**** Details ****\n')
tdf.reset_index(inplace=True)
tdf['Units'] = tdf['Volume']
tdf['Symbol'] = tdf.apply(lambda row: row['Market'][:-3].upper(), axis=1)
tdf['Cost'] = tdf['Total'] * 1000
tdf['Average'] = tdf['Average'] * 1000
tdf['Desc'] = tdf['Symbol']
tdf.drop(['Total', 'Volume', 'Market'], axis=1, inplace=True)
print(tdf)

# Generate fund excel
fname = basecryptocostdet+nowstr+'.xlsx'
tdf.to_excel(fname, index=False, engine='xlsxwriter')
print("\nWritten %d rows to file %s" % (len(tdf),fname))

print('\n**** Aggregate ****\n')
tdf = tdf.groupby('Symbol').agg(np.sum)
tdf['Average'] = tdf['Cost']/tdf['Units']
print(tdf)
print(tdf.index.values)
tot = tdf['Cost'].sum()
print('\nTotal: %.5f lacs' % (tot/100000))

print('='*80)
